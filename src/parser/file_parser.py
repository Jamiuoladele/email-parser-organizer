import csv
import xml.etree.ElementTree as ET
from openpyxl import load_workbook
from PyPDF2  import  PdfReader


def parse_csv(file_path):
    emails = []

    with open(file_path, newline='') as csvfile:
        reader = csv.DictReader(csvfile)
        for row in reader:
            emails.append({
                "sender": row.get("sender", ""),
                "subject": row.get("subject", ""),
                "body": row.get("body", "")
            })

    return emails


def parse_xml(file_path):
    emails = []

    tree = ET.parse(file_path)
    root = tree.getroot()

    for email in root.findall("email"):
        emails.append({
            "sender": email.findtext("sender", ""),
            "subject": email.findtext("subject", ""),
            "body": email.findtext("body", "")
        })

    return emails


# 🔥 NEW FUNCTION
def parse_excel(file_path):
    emails = []

    workbook = load_workbook(file_path)
    sheet = workbook.active

    # Skip header row
    for row in sheet.iter_rows(min_row=2, values_only=True):
        sender, subject, body = row

        emails.append({
            "sender": sender or "",
            "subject": subject or "",
            "body": body or ""
        })

    return emails
def parse_pdf(file_path):
    emails = []

    reader = PdfReader(file_path)
    text = ""

    for page in reader.pages:
        text += page.extract_text() + "\n"

    emails.append({
        "sender": "pdf@upload",
        "subject": "PDF Document",
        "body": text
    })

    return emails


def parse_file(file_path):
    if file_path.endswith(".csv"):
        return parse_csv(file_path)

    elif file_path.endswith(".xml"):
        return parse_xml(file_path)

    elif file_path.endswith(".xlsx"):
        return parse_excel(file_path)
    elif file_path.endswith(".pdf"):
        return parse_pdf(file_path)

    else:
        raise ValueError("Unsupported file format")